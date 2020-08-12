import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import OneHotEncoder
from sklearn.compose import ColumnTransformer
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, f1_score
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

coord_file = 'Coordinates.xlsx'
report_file = 'Report Territory Management.xlsx'
output_file = 'Report Territory Management Updated.xlsx'

print('Step 1 of 9: Prepare coordinates')

df_coor = pd.read_excel(coord_file)
df_coor.columns = ['SWE_Store_Key','Latitude','Longitude']
# cleansing from invalid coordinates
df_coor = df_coor[df_coor['Latitude']!=0]
df_coor = df_coor[ ( (df_coor['Latitude']>40) & (df_coor['Latitude']<82) ) ]
df_coor = df_coor[ ( ( (df_coor['Longitude']>=10) & (df_coor['Longitude']<180) ) | \
    ( (df_coor['Longitude']>=-180) & (df_coor['Longitude']<-160) ) ) ]
# check if outlets are duplicated
if df_coor.SWE_Store_Key.value_counts().values[0] > 1:
    print('Duplicated outlets were found in a file with coordinates')
    assert False

print('Step 2 of 9: Prepare territories')

df_terr = pd.read_excel(report_file,skiprows=1)

# rename fields
df_terr.columns=['Region','Distrib','Office','FFDSL','TSE_MTDE','Level_Torg_Region1','Level_Torg_Region2', \
'Filial_Name','Filial_Ship_To','Chain_Type','Chain_Name','Chain_Id','Chain_Chain_Tier_MWC','Chain_Chain_Sub_Tier_MWC', \
'SWE_Store_Key','Store_Status','Outlet_Name','Channel_Name_2018','Outlet_Type_2018','Trade_Structure','From_Dc', \
'Segment_MWC_Segment_Name','Cluster_MWC','Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4','Kladr_level_5', \
'LSV_WWY','LSV_CHOCO','LSV_MWC','Covering_Outlet_id','General_Duplicate','Ship_To_Visited','Filial_Visited', \
'Ship_to_Name_TO_BE','Region_loaded_RSS','Ship_to_TO_BE_Name_loaded_RSS','SHIP_TO_RSS', \
'Ship_to_Code_TO_BE','DC','Changed','Change_Period']

df_codes = pd.DataFrame(data=df_terr['SWE_Store_Key'],columns=['SWE_Store_Key'])

# do not take unused columns
df_terr = df_terr[['SWE_Store_Key','Region','Distrib','Office','FFDSL','TSE_MTDE','Level_Torg_Region1','Level_Torg_Region2',
'Filial_Name','Filial_Ship_To','Chain_Type','Chain_Name','Chain_Id','Chain_Chain_Tier_MWC','Chain_Chain_Sub_Tier_MWC',
'Channel_Name_2018','Outlet_Type_2018','Trade_Structure','From_Dc','Segment_MWC_Segment_Name','Cluster_MWC',
'Covering_Outlet_id','General_Duplicate','SHIP_TO_RSS','Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4','Kladr_level_5']]
# Remove outlet-duplicates and associated fields
df_terr = df_terr[df_terr['General_Duplicate']!='Дубликат']
df_terr.drop(['Covering_Outlet_id','General_Duplicate'],axis=1,inplace=True)

print('Step 3 of 9: Merge territories with coordinates and start preprocessing')

df = pd.merge(df_terr, df_coor, on='SWE_Store_Key',how='left')
del df_terr
del df_coor

# SHIP_TO_RSS should be int value
df.loc[~df['SHIP_TO_RSS'].isna(),'SHIP_TO_RSS'] = df.loc[~df['SHIP_TO_RSS'].isna()]['SHIP_TO_RSS'].apply(lambda x: int(x[:9]))
# Chain_Name duplicates Chain_Id
df.drop(['Chain_Name'],axis=1,inplace=True)

df['Trade_Structure'].fillna('Не известно',inplace=True)
df.Trade_Structure.value_counts()

df['isTrain'] = ~ df['SHIP_TO_RSS'].isna()
df['isCoord'] = ~( (df['Latitude'].isna()) | (df['Longitude'].isna()) )

# If there is no coordinates for an outlet, get average coordinates from the minimal known regional division
print('Step 4 of 9: Restore coordinates')

kladr_lat_grouped = df[df['isCoord']==1].groupby(['Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4']).Latitude.mean()
kladr_lon_grouped = df[df['isCoord']==1].groupby(['Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4']).Longitude.mean()

def get_avg_latitude(row):
    try:
        return kladr_lat_grouped[row['Kladr_level_1']][row['Kladr_level_2']][row['Kladr_level_3']][row['Kladr_level_4']]
    except:
        try:
            return kladr_lat_grouped[row['Kladr_level_1']][row['Kladr_level_2']][row['Kladr_level_3']].mean()
        except:
            try:
                return kladr_lat_grouped[row['Kladr_level_1']][row['Kladr_level_2']].mean()
            except:
                try:
                    return kladr_lat_grouped[row['Kladr_level_1']].mean()
                except:
                    return 0

def get_avg_longitude(row):
    try:
        return kladr_lon_grouped[row['Kladr_level_1']][row['Kladr_level_2']][row['Kladr_level_3']][row['Kladr_level_4']]
    except:
        try:
            return kladr_lon_grouped[row['Kladr_level_1']][row['Kladr_level_2']][row['Kladr_level_3']].mean()
        except:
            try:
                return kladr_lon_grouped[row['Kladr_level_1']][row['Kladr_level_2']].mean()
            except:
                try:
                    return kladr_lon_grouped[row['Kladr_level_1']].mean()
                except:
                    return 0

df.loc[df['isCoord']==0,'Latitude'] = \
    df.loc[df['isCoord']==0][['SWE_Store_Key','Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4']].apply( \
        get_avg_latitude,axis=1)

df.loc[df['isCoord']==0,'Longitude'] = \
    df.loc[df['isCoord']==0][['SWE_Store_Key','Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4']].apply( \
        get_avg_longitude,axis=1)

# Address classificator not required anymore
df.drop(['Kladr_level_1','Kladr_level_2','Kladr_level_3','Kladr_level_4','Kladr_level_5'],axis=1,inplace=True)

# remove outlets with null coordinates
df = df[~((df['Latitude']==0) | (df['Longitude']==0))]

df.set_index('SWE_Store_Key',inplace=True)
df['From_Dc'] = df['From_Dc'].astype(int)

# Save DataFrame for future usage
df_init = df.copy()

# Train model
target = ['SHIP_TO_RSS']
service = ['isTrain','isCoord']
features = [column for column in df.columns if column not in target and column not in service]
X = df[df['isTrain']==True][features]
y = df[df['isTrain']==True][target]
y['SHIP_TO_RSS'] = y['SHIP_TO_RSS'].astype(int)

cat_features = X.select_dtypes(include=['object']).columns  # Categorical
num_features = X.select_dtypes(exclude=['object']).columns  # Numeric

cat_features, num_features

for col in cat_features:
    X[col]=X[col].astype(str)

X_train, X_valid, y_train, y_valid = train_test_split(X, y, test_size=0.3, random_state=42)

## Pipeline for train
print('Step 5 of 9: Training the model')

cat_pipe = Pipeline([
    ('imputer', SimpleImputer(strategy='constant', fill_value='missing')),
    ('onehot', OneHotEncoder(handle_unknown='ignore'))
])
num_pipe = Pipeline([
    ('imputer', SimpleImputer(strategy='mean'))
])

preprocessing = ColumnTransformer(
    [('cat', cat_pipe, cat_features),
     ('num', num_pipe, num_features)
    ])

rf = Pipeline([
    ('preprocess', preprocessing),
    ('classifier', RandomForestClassifier(n_estimators=10, random_state=42, n_jobs=-1))
])

rf.fit(X_train, y_train.values.ravel())

def print_logisitc_metrics(y_true, y_pred):
    acc = accuracy_score(y_true, y_pred)
    f1 = f1_score(y_true, y_pred,average='micro')
    print(f'Test dataset accuracy = {acc:.5f}, f1-score = {f1:.5f}')

y_pred = rf.predict(X_valid)

print_logisitc_metrics(y_valid, y_pred)

# All-data training

print('Step 6 of 9: Train model with full data and predict values')

rf.fit(X, y.values.ravel())

print('Step 7 of 9: Define top 3 classes')
# Define top 3 classes for each outlet without an answer

X_pred = df[df['isTrain']==False][features]
y_pred_proba = rf.predict_proba(X_pred)
df_proba = pd.DataFrame(data=y_pred_proba,columns=rf.classes_)

def get_three_max(row):
    ser = pd.Series(data=row.values, index=rf.classes_)
    ser.sort_values(inplace=True,ascending=False)
    return ser[0:3].index[0],ser[0:3].values[0],ser[0:3].index[1],ser[0:3].values[1],ser[0:3].index[2],ser[0:3].values[2]

df_proba['top_1_class'], df_proba['top_1_proba'],df_proba['top_2_class'], df_proba['top_2_proba'], \
    df_proba['top_3_class'], df_proba['top_3_proba'] = zip(*df_proba.apply(get_three_max,axis=1))

print('Step 8 of 9: Prepare output file')

X_pred.reset_index(inplace=True)
df_concat = pd.concat([X_pred['SWE_Store_Key'], df_proba.loc[:,'top_1_class':]], axis=1,join='inner')

df_info = df_codes.merge(right=df_concat,how='left',on='SWE_Store_Key')

del df_codes
del df_concat

workbook = openpyxl.load_workbook(report_file)
worksheet = workbook['Sheet1']

rows = dataframe_to_rows(df_info, index=False, header=True)

start_row = 2
start_col = 46
for r_idx, row in enumerate(rows, start_row):
    for c_idx, value in enumerate(row, start_col):
         worksheet.cell(row=r_idx, column=c_idx, value=value)
    if type(row[2])==float and not pd.isnull(row[2]):
        proba = float(row[2])
        if proba >= 0.75:
            worksheet.cell(row=r_idx, column=39, value=row[1])
            worksheet.cell(row=r_idx, column=39).fill = PatternFill("solid", fgColor="00D328")
        elif proba >= 0.5:
            worksheet.cell(row=r_idx, column=39, value=row[1])
            worksheet.cell(row=r_idx, column=39).fill = PatternFill("solid", fgColor="F9F405")
        else:
            worksheet.cell(row=r_idx, column=39, value=row[1])
            worksheet.cell(row=r_idx, column=39).fill = PatternFill("solid", fgColor="FCB09F")

print('Step 9 of 9: Saving output file')
workbook.save(output_file)

print('The new Territory Management report has been saved as', output_file)
